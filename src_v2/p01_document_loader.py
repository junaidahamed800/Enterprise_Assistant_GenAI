import os
from langchain_community.document_loaders import (
    TextLoader,
    PyPDFLoader,
    Docx2txtLoader,
)
from langchain_core.documents import Document


def load_single_file(filepath: str) -> list[Document]:
    """
    Load a single file based on its extension.
    Returns a list of LangChain Document objects.

    WHY: Each file type needs a different loader because
    they store content differently internally.
    - TXT  → plain text, read directly
    - PDF  → binary format, need PyPDF to extract text per page
    - DOCX → XML-based zip, need Docx2txt to extract paragraphs
    - XLSX → spreadsheet, we read each sheet row by row as text
    """

    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".txt":
        loader = TextLoader(filepath, encoding="utf-8")
        return loader.load()

    elif ext == ".pdf":
        loader = PyPDFLoader(filepath)
        return loader.load()

    elif ext == ".docx":
        loader = Docx2txtLoader(filepath)
        return loader.load()

    elif ext == ".xlsx":
        # LangChain doesn't have a great built-in xlsx loader,
        # so we read it manually using openpyxl and wrap content
        # in a Document object — same format LangChain expects.
        return _load_excel(filepath)

    else:
        print(f"[SKIP] Unsupported file type: {filepath}")
        return []


def _load_excel(filepath: str) -> list[Document]:
    """
    Reads each sheet of an Excel file.
    Each row becomes a line of text.
    The whole sheet is one Document.

    WHY: We convert rows to text so the embedding model
    (which only understands text) can work with Excel data.
    """
    import openpyxl

    workbook = openpyxl.load_workbook(filepath, data_only=True)
    documents = []

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        lines = []

        for row in sheet.iter_rows(values_only=True):
            # Filter out completely empty rows
            row_text = " | ".join(
                str(cell) for cell in row if cell is not None
            )
            if row_text.strip():
                lines.append(row_text)

        if lines:
            content = f"Sheet: {sheet_name}\n" + "\n".join(lines)
            documents.append(
                Document(
                    page_content=content,
                    metadata={"source": filepath, "sheet": sheet_name}
                )
            )

    return documents


def load_documents(data_dir: str = "data") -> list[Document]:
    """
    Scans the data/ folder and loads ALL supported files.
    Supported: .txt, .pdf, .docx, .xlsx

    WHY: By scanning a folder instead of hardcoding filenames,
    you can drop any new file into data/ and it gets picked up
    automatically — no code changes needed.
    """

    supported_extensions = {".txt", ".pdf", ".docx", ".xlsx"}
    documents = []

    if not os.path.exists(data_dir):
        print(f"[WARNING] data directory '{data_dir}' not found.")
        return documents

    for filename in os.listdir(data_dir):
        ext = os.path.splitext(filename)[1].lower()

        if ext not in supported_extensions:
            continue

        filepath = os.path.join(data_dir, filename)
        print(f"[LOADING] {filename}")

        try:
            docs = load_single_file(filepath)
            documents.extend(docs)
            print(f"  → {len(docs)} document(s) loaded")

        except Exception as e:
            print(f"  → [ERROR] Could not load {filename}: {e}")

    print(f"\nTotal documents loaded: {len(documents)}")
    return documents
